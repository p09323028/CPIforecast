"""產生台灣 CPI 預測報表（xlsx）— 仿 USDA TB-1957 月報格式。

從 raw_cpi.csv 與每類別的 rolling_yoy.csv 即時計算各欄位，輸出有格式
的 xlsx，使用者可直接下載作為新聞稿或內部報告。
"""
from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import storage
from .categories import DISPLAY_ZH, english_names

log = logging.getLogger(__name__)

# 子項目（隸屬於上方的大類）在「品項」欄縮排一個全形空格，呈現層級：
#   肉類 → 豬肉 / 牛肉 / 雞肉、蛋類 → 雞蛋、乳類 → 鮮乳
SUB_ITEMS = {"pork", "beef", "poultry", "egg", "milk"}
_INDENT = "　"  # 全形空格


def _mom(s: pd.Series, date: pd.Timestamp) -> Optional[float]:
    """本月對上月的 % 變動率。"""
    prev = date - pd.DateOffset(months=1)
    if date not in s.index or prev not in s.index:
        return None
    v0, v1 = s[prev], s[date]
    if v0 == 0 or pd.isna(v0) or pd.isna(v1):
        return None
    return float((v1 / v0 - 1) * 100)


def _yoy(s: pd.Series, date: pd.Timestamp) -> Optional[float]:
    """本月對去年同月的 % 變動率。"""
    prev = date - pd.DateOffset(years=1)
    if date not in s.index or prev not in s.index:
        return None
    v0, v1 = s[prev], s[date]
    if v0 == 0 or pd.isna(v0) or pd.isna(v1):
        return None
    return float((v1 / v0 - 1) * 100)


def _ytd_avg(s: pd.Series, date: pd.Timestamp) -> Optional[float]:
    """年初至今平均 % 變動：本年 Jan~{date} 平均 vs 去年同期平均。"""
    year = date.year
    cur = s[(s.index.year == year) & (s.index <= date)]
    if cur.empty:
        return None
    prev_idx = [d - pd.DateOffset(years=1) for d in cur.index]
    prev = s.reindex(prev_idx).dropna()
    if prev.empty:
        return None
    cur_mean = cur.mean()
    prev_mean = prev.mean()
    if prev_mean == 0:
        return None
    return float((cur_mean / prev_mean - 1) * 100)


def _annual(s: pd.Series, year: int) -> Optional[float]:
    """指定年度的年平均 % 變動率（vs. 前年年平均）。需兩年皆 12 個月完整。"""
    cur = s[s.index.year == year]
    prev = s[s.index.year == year - 1]
    if len(cur) < 12 or len(prev) < 12:
        return None
    return float((cur.mean() / prev.mean() - 1) * 100)


def _hist_avg(s: pd.Series) -> Optional[float]:
    """所有完整年度的年變動率算術平均。"""
    years = sorted({int(y) for y in s.index.year})
    changes = [c for c in (_annual(s, y) for y in years) if c is not None]
    if not changes:
        return None
    return float(sum(changes) / len(changes))


def _read_latest_rolling(
    base_dir: str | Path, run_id: str, cat: str, data_end: pd.Timestamp
) -> tuple[Optional[float], Optional[float], Optional[float]]:
    """讀該類別 rolling_yoy.csv，回傳 effective_end == data_end 那筆的
    (lower_95, median, upper_95)。"""
    path = storage.category_dir(base_dir, run_id, cat) / storage.ROLLING_YOY_CSV
    if not path.exists():
        return None, None, None
    df = pd.read_csv(path)
    end_str = data_end.strftime("%Y-%m-%d")
    match = df[df["effective_end"] == end_str]
    if match.empty:
        match = df.tail(1)
    if match.empty:
        return None, None, None
    row = match.iloc[-1]
    return (
        _safe_float(row.get("lower_95")),
        _safe_float(row.get("median")),
        _safe_float(row.get("upper_95")),
    )


def _safe_float(v) -> Optional[float]:
    try:
        if v is None or pd.isna(v):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def build_report_df(
    base_dir: str | Path, run_id: str, manifest: dict
) -> pd.DataFrame:
    """建構報表的 DataFrame（14 列 × N 欄）。"""
    data_end = pd.to_datetime(manifest["data_end_date"])
    forecast_year = data_end.year

    raw_path = storage.run_dir(base_dir, run_id) / storage.RAW_CSV
    raw = pd.read_csv(raw_path, index_col="date", parse_dates=True).asfreq("MS")

    rows: list[dict] = []
    for en in english_names():
        s = raw[en].dropna()
        lo, mid, hi = _read_latest_rolling(base_dir, run_id, en, data_end)
        name = DISPLAY_ZH[en]
        if en in SUB_ITEMS:
            name = _INDENT + name
        rows.append(
            {
                "品項": name,
                f"月變動率% ({data_end.month}月)": _mom(s, data_end),
                f"年變動率% ({data_end.month}月)": _yoy(s, data_end),
                f"年初至今平均變動率% ({forecast_year})": _ytd_avg(s, data_end),
                "2023年": _annual(s, 2023),
                "2024年": _annual(s, 2024),
                "2025年": _annual(s, 2025),
                f"歷史平均% ({s.index.min().year}-{s.index.max().year})": _hist_avg(s),
                f"預測{forecast_year}年下限%": lo,
                f"預測{forecast_year}年中位數%": mid,
                f"預測{forecast_year}年上限%": hi,
            }
        )
    return pd.DataFrame(rows)


def write_xlsx(
    base_dir: str | Path,
    run_id: str,
    manifest: dict,
    out: io.BytesIO | Path,
) -> None:
    """產生格式化的 xlsx 報表寫入 `out`（BytesIO 或檔案路徑）。"""
    df = build_report_df(base_dir, run_id, manifest)
    data_end = pd.to_datetime(manifest["data_end_date"])
    forecast_year = data_end.year

    wb = Workbook()
    ws = wb.active
    ws.title = f"{forecast_year}年{data_end.month}月CPI預測"

    title = (
        f"台灣食物類 CPI 變動率與預測表（資料截至 "
        f"{data_end.strftime('%Y年%m月')}）"
    )
    n_cols = len(df.columns)

    # === 標題列 ===
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Microsoft JhengHei", size=14, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # === 表頭 ===
    header_fill = PatternFill("solid", fgColor="E0E7FF")
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True)
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for i, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=i, value=col)
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.fill = header_fill
        cell.border = border
    ws.row_dimensions[2].height = 50

    # === 資料列 ===
    body_font = Font(name="Microsoft JhengHei", size=11)
    name_font = Font(name="Microsoft JhengHei", size=11, bold=True)
    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if c_idx == 1:
                cell.value = val
                cell.font = name_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    cell.value = "—"
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                else:
                    cell.value = float(val)
                    cell.number_format = "+0.00;-0.00;0.00"
                    cell.alignment = Alignment(
                        horizontal="right", vertical="center"
                    )
                cell.font = body_font
            cell.border = border
        ws.row_dimensions[r_idx].height = 22

    # === 歷史欄與預測欄之間的虛線分隔 ===
    # 在第一個「預測…」欄的左側畫虛線，把歷史/實際欄與預測欄區隔開。
    sep_col = next(
        (i for i, col in enumerate(df.columns, start=1) if str(col).startswith("預測")),
        None,
    )
    if sep_col is not None:
        dashed = Side(style="mediumDashed", color="475569")
        last_row = 2 + len(df)  # 表頭(第 2 列) + 14 列資料
        for r in range(2, last_row + 1):
            cell = ws.cell(row=r, column=sep_col)
            prev = cell.border
            cell.border = Border(
                left=dashed,
                right=prev.right,
                top=prev.top,
                bottom=prev.bottom,
            )

    # === 欄寬 ===
    ws.column_dimensions["A"].width = 14
    for i in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # === 註腳 ===
    notes_row = 3 + len(df) + 1
    notes = [
        "說明：",
        "1. 資料來源：行政院主計總處 dataset 6019（消費者物價指數 CPI 月資料）。",
        "2. 模型方法：USDA TB-1957 SARIMA 自動配適 + 蒙地卡羅模擬（10,000 條路徑）。",
        f"3. 預測{forecast_year}年區間以「資料截至 {data_end.strftime('%Y-%m')}」之模型推估；當年完整 12 個月的年平均變動率分位數（2.5% / 50% / 97.5%）。",
        "4. 各欄位定義：",
        "   - 月變動率：本月指數 / 上月指數 − 1（%）",
        "   - 年變動率：本月指數 / 去年同月指數 − 1（%）",
        "   - 年初至今平均變動率：本年 1 月至本月平均 / 去年同期平均 − 1（%）",
        "   - 年度變動率：當年平均 / 前年平均 − 1（%），需兩年皆有 12 個月完整資料",
        "   - 歷史平均變動率：所有完整年度之年變動率的算術平均",
        "5. 14 項類別對應主計總處公布之消費者物價基本分類項目。",
    ]
    notes_font = Font(name="Microsoft JhengHei", size=9, color="475569")
    for i, text in enumerate(notes):
        cell = ws.cell(row=notes_row + i, column=1, value=text)
        cell.font = notes_font
        ws.merge_cells(
            start_row=notes_row + i,
            start_column=1,
            end_row=notes_row + i,
            end_column=n_cols,
        )

    # 凍結首列 + 表頭
    ws.freeze_panes = "B3"

    if isinstance(out, (str, Path)):
        wb.save(str(out))
    else:
        wb.save(out)
