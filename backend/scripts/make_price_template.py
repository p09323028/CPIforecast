"""產生「實際價格」預測用的 Excel 填寫範本。

非 CPI 指數資料（例：雞蛋產地價格、甘藍批發價格）也想用與 CPI 平台相同的
SARIMA + 蒙地卡羅流程預測。本腳本產生一個讓使用者填值的 xlsx：
  - 工作表「說明」：填寫規則
  - 工作表「價格資料」：A 欄月份（已預填 2014-06~2026-05），B 欄起每欄一個品項

填好後交給 forecast_prices.py 跑模型（輸出比照 CPI：月度/年 YoY 分位數）。

用法：
    python scripts/make_price_template.py
產生：backend/data/price_inputs/price_template.xlsx
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parents[1] / "data" / "price_inputs"
OUT_PATH = OUT_DIR / "price_template.xlsx"

# 預填月份：對齊模型 12 年訓練窗（最近 144 個月，至 2026-05）
START = "2014-06-01"
END = "2026-05-01"

# 範例品項欄（標題可自行修改／增刪），含單位範例
EXAMPLE_ITEMS = [
    "雞蛋產地價格(元/台斤)",
    "甘藍批發價格(元/公斤)",
    "（品項3：自行命名）",
    "（品項4：自行命名）",
]

HEADER_FILL = PatternFill("solid", fgColor="E0E7FF")
HEADER_FONT = Font(name="Microsoft JhengHei", size=11, bold=True)
NOTE_FONT = Font(name="Microsoft JhengHei", size=11)
TITLE_FONT = Font(name="Microsoft JhengHei", size=14, bold=True)


def _build_instructions(ws) -> None:
    ws["A1"] = "實際價格 — 預測資料填寫範本"
    ws["A1"].font = TITLE_FONT
    notes = [
        "",
        "用途：填入「實際價格」資料（非指數），例如雞蛋產地價格、甘藍批發價格，",
        "      之後會用與 CPI 平台相同的模型（SARIMA 自動配適 + 10,000 條蒙地卡羅模擬）跑預測。",
        "",
        "填寫步驟：",
        "  1. 切到「價格資料」工作表。",
        "  2. A 欄「月份」已預先填好（2014-06 ~ 2026-05）。從你資料最早的月份開始填值，",
        "     開頭沒有資料的月份請「整列刪除」。",
        "  3. B 欄起，每一欄是一個品項。直接把對應月份的價格數字填進去。",
        "  4. 標題列（第 1 列）可改成你要的品項名稱，建議含單位，例：雞蛋產地價格(元/台斤)。",
        "  5. 要新增品項：在右邊空白欄繼續加標題與數值即可，沒有欄數上限。",
        "",
        "重要規則：",
        "  • 頻率為「每月一筆」。原始若是每日／每週資料，請先彙整成『月平均』再填。",
        "  • 資料必須『連續、不可跳月』（SARIMA 季節模型不允許中間缺月）。",
        "  • 數值就是實際價格（可含小數），不要填百分比、文字或單位符號。",
        "  • 末端對齊到你最新一個『完整月』即可（未滿月的當月先不要填）。",
        "  • 建議至少 60 個月（5 年）；本模型用最近 144 個月（12 年）訓練，資料越長越準。",
        "",
        "填好存檔後，把檔案告訴我，我會跑出（比照 CPI 平台）：",
        "  - 未來 18 個月的價格中位數與 95% 預測區間",
        "  - 年平均的年變動率（YoY%）分位數",
        "  - 每品項一組與 CPI 相同格式的 CSV／圖表資料",
    ]
    for i, text in enumerate(notes, start=2):
        c = ws.cell(row=i, column=1, value=text)
        c.font = NOTE_FONT
        c.alignment = Alignment(vertical="center")
    ws.column_dimensions["A"].width = 100


def _build_data_sheet(ws) -> None:
    months = pd.date_range(START, END, freq="MS")

    # 表頭
    ws.cell(row=1, column=1, value="月份")
    for j, name in enumerate(EXAMPLE_ITEMS, start=2):
        ws.cell(row=1, column=j, value=name)
    for col in range(1, 1 + len(EXAMPLE_ITEMS) + 1):
        c = ws.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 月份欄（真實日期，顯示 yyyy-mm）
    for i, d in enumerate(months, start=2):
        c = ws.cell(row=i, column=1, value=d.to_pydatetime())
        c.number_format = "yyyy-mm"
        c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    for j in range(2, 2 + len(EXAMPLE_ITEMS)):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 32


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_notes = wb.active
    ws_notes.title = "說明"
    _build_instructions(ws_notes)
    ws_data = wb.create_sheet("價格資料")
    _build_data_sheet(ws_data)
    wb.save(OUT_PATH)
    print(f"範本已產生：{OUT_PATH}")


if __name__ == "__main__":
    main()
